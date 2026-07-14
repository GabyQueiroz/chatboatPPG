from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FAQ_PATH = ROOT / "static" / "FAQ.md"
DOCS_DIR = ROOT / "docs"
OUTPUT_PATH = ROOT / "base_perguntas_respostas_ppgd.xlsx"


def clean_heading(text: str) -> str:
    return re.sub(r"^\d+\.\s*", "", text).strip()


def parse_faq():
    text = FAQ_PATH.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^###\s+(.+)$", text, flags=re.MULTILINE))
    entries = []

    for index, match in enumerate(matches):
        question = clean_heading(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()

        source_match = re.search(r"Fonte[s]?:\s*(.+)", block, flags=re.IGNORECASE)
        source = source_match.group(1).strip() if source_match else "FAQ/documentos"

        answer_block = re.split(r"Fonte[s]?:", block, maxsplit=1, flags=re.IGNORECASE)[0]
        answer_lines = []
        for line in answer_block.splitlines():
            line = line.strip()
            if not line or line.startswith("##"):
                continue
            answer_lines.append(line)
        answer = " ".join(answer_lines).strip()

        if question and answer:
            entries.append((question, answer, source))

    return entries


def category_for(question: str) -> str:
    folded = question.lower()
    if any(term in folded for term in ["e-mail", "email", "instagram", "publicações", "contato"]):
        return "contato_links"
    if any(term in folded for term in ["disciplina", "crédito", "grade", "curricular", "odm"]):
        return "curriculo_disciplinas"
    if any(term in folded for term in ["prazo", "qualificação", "suficiência", "língua"]):
        return "prazos_requisitos"
    if any(term in folded for term in ["defesa", "dissertação", "termo", "ficha", "cadastro"]):
        return "defesa_documentos"
    if any(term in folded for term in ["fomento", "laboratório", "legaltech", "red"]):
        return "normas_procedimentos"
    return "programa"


def base_intent(question: str) -> str:
    text = question.lower()
    text = re.sub(r"[^a-z0-9áéíóúàâêôãõç]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:48] or "pergunta"


def variations(question: str):
    q = question.rstrip("?")
    vars_ = [question]

    replacements = [
        ("Qual é ", "Me informe "),
        ("Quais são ", "Liste "),
        ("Como ", "De que forma "),
        ("O que é ", "Explique "),
        ("O que são ", "Explique "),
        ("Para que serve ", "Qual a finalidade de "),
        ("É possível ", "Posso "),
        ("A ", ""),
        ("O ", ""),
    ]

    for old, new in replacements:
        if q.startswith(old):
            vars_.append(new + q[len(old):] + "?")
            break

    vars_.append(f"Gostaria de saber: {q.lower()}?")
    vars_.append(f"Você pode me responder {q.lower()}?")

    # Domain-specific synonyms.
    more = []
    for item in vars_:
        more.append(item)
        more.append(item.replace("Mestrado Profissional em Direito", "PPGD"))
        more.append(item.replace("Programa", "curso").replace("programa", "curso"))

    deduped = []
    seen = set()
    for item in more:
        item = re.sub(r"\s+", " ", item).strip()
        if item and item not in seen:
            seen.add(item)
            deduped.append(item)

    return deduped[:4]


def fallback_rows():
    default = (
        "Desculpe, não tenho informações suficientes para responder a essa pergunta. "
        "Posso tentar ajudar com informações sobre prazos, disciplinas, créditos, defesa, "
        "suficiência, atividades complementares, documentos, contatos ou links do Programa."
    )
    return [
        {
            "id": "fallback_mensalidade_01",
            "categoria": "fallback",
            "intencao": "fora_da_base",
            "pergunta_base": "Qual é o valor da mensalidade?",
            "pergunta_variacao": "Quanto custa o mestrado?",
            "resposta_esperada": default,
            "fonte": "Sem informação nos documentos analisados",
            "tipo_resposta": "fallback",
            "pergunta_semelhante_sugerida": "Onde acompanho publicações do mestrado?",
            "resposta_quando_nao_entender": default,
        },
        {
            "id": "fallback_edital_01",
            "categoria": "fallback",
            "intencao": "edital_atual",
            "pergunta_base": "Quando abre inscrição para a próxima turma?",
            "pergunta_variacao": "Qual a data do próximo edital?",
            "resposta_esperada": (
                "Desculpe, não tenho informações suficientes para responder a essa pergunta com segurança. "
                "Para editais e notícias, consulte as publicações oficiais do Mestrado."
            ),
            "fonte": "Sem calendário atual nos documentos analisados",
            "tipo_resposta": "fallback",
            "pergunta_semelhante_sugerida": "Onde acompanho publicações do mestrado?",
            "resposta_quando_nao_entender": default,
        },
        {
            "id": "fallback_ambigua_01",
            "categoria": "fallback",
            "intencao": "pergunta_ambigua",
            "pergunta_base": "Qual é o prazo?",
            "pergunta_variacao": "E isso, como funciona?",
            "resposta_esperada": (
                "Você poderia reformular a pergunta informando o tema? Posso ajudar com prazos de conclusão, "
                "qualificação, defesa, disciplinas, créditos, suficiência ou atividades complementares."
            ),
            "fonte": "Resposta padrão para baixa confiança",
            "tipo_resposta": "fallback",
            "pergunta_semelhante_sugerida": "Qual o prazo para entregar a versão final após a defesa?",
            "resposta_quando_nao_entender": default,
        },
    ]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="0F6B5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook():
    entries = parse_faq()
    rows = []
    default_fallback = (
        "Desculpe, não tenho informações suficientes para responder a essa pergunta. "
        "Posso tentar ajudar com informações sobre prazos, disciplinas, créditos, defesa, "
        "suficiência, atividades complementares, documentos, contatos ou links do Programa."
    )

    for entry_index, (question, answer, source) in enumerate(entries, start=1):
        intent = base_intent(question)
        category = category_for(question)
        for var_index, variation in enumerate(variations(question), start=1):
            rows.append(
                {
                    "id": f"q{entry_index:03d}_{var_index:02d}",
                    "categoria": category,
                    "intencao": intent,
                    "pergunta_base": question,
                    "pergunta_variacao": variation,
                    "resposta_esperada": answer,
                    "fonte": source,
                    "tipo_resposta": "direta" if category == "contato_links" else "rag",
                    "pergunta_semelhante_sugerida": question,
                    "resposta_quando_nao_entender": default_fallback,
                }
            )

    rows.extend(fallback_rows())

    wb = Workbook()
    ws = wb.active
    ws.title = "Perguntas_Respostas"
    headers = [
        "id",
        "categoria",
        "intencao",
        "pergunta_base",
        "pergunta_variacao",
        "resposta_esperada",
        "fonte",
        "tipo_resposta",
        "pergunta_semelhante_sugerida",
        "resposta_quando_nao_entender",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    widths = [20, 22, 38, 52, 58, 95, 54, 16, 52, 95]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    style_sheet(ws)
    table = Table(displayName="TabelaPerguntasRespostas", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)

    fb = wb.create_sheet("Fallback")
    fb.append(["cenario", "gatilho", "resposta_padrao", "acao_recomendada", "exemplo_pergunta_semelhante"])
    fb_data = [
        [
            "Pergunta fora da base",
            "Nenhum contexto recuperado com segurança",
            default_fallback,
            "Sugerir temas cobertos pelo acervo.",
            "Quais são os requisitos para fazer o Exame de Qualificação?",
        ],
        [
            "Pergunta ambígua",
            "Pergunta curta ou sem tema claro",
            "Você poderia reformular a pergunta informando o tema?",
            "Pedir esclarecimento antes de responder.",
            "Qual o prazo máximo para conclusão do mestrado?",
        ],
        [
            "Pergunta sobre link",
            "link, site, Instagram, página, acesso",
            "Retornar a URL exata encontrada no contexto.",
            "Não alterar caracteres da URL.",
            "Qual é o Instagram do mestrado?",
        ],
        [
            "Pergunta sobre contato",
            "email, e-mail, telefone, ramal",
            "Retornar exatamente o contato encontrado no contexto.",
            "Usar extração direta para evitar alteração pelo modelo.",
            "Qual é o e-mail do mestrado?",
        ],
    ]
    for row in fb_data:
        fb.append(row)
    for idx, width in enumerate([30, 44, 90, 62, 54], start=1):
        fb.column_dimensions[get_column_letter(idx)].width = width
    style_sheet(fb)

    sources = wb.create_sheet("Fontes")
    sources.append(["arquivo", "tipo", "observacao"])
    files = sorted(p for p in DOCS_DIR.rglob("*") if p.is_file())
    for p in files:
        sources.append([p.relative_to(ROOT).as_posix(), p.suffix.lower().lstrip("."), "Documento considerado"])
    sources.column_dimensions["A"].width = 110
    sources.column_dimensions["B"].width = 14
    sources.column_dimensions["C"].width = 34
    style_sheet(sources)

    summary = wb.create_sheet("Resumo")
    summary.append(["campo", "valor"])
    summary.append(["total_perguntas_variacoes", len(rows)])
    summary.append(["total_perguntas_base", len(entries)])
    summary.append(["total_fontes_docs", len(files)])
    summary.append(["arquivo_origem_faq", FAQ_PATH.relative_to(ROOT).as_posix()])
    summary.append(["observacao", "Cada linha da aba Perguntas_Respostas é uma pergunta ou variação com resposta esperada."])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 90
    style_sheet(summary)

    wb.save(OUTPUT_PATH)

    check = load_workbook(OUTPUT_PATH, read_only=True)
    total = check["Perguntas_Respostas"].max_row - 1
    print(f"arquivo={OUTPUT_PATH}")
    print(f"perguntas_variacoes={total}")
    print(f"fontes={len(files)}")
    if total <= 100:
        raise RuntimeError("A planilha precisa ter mais de 100 perguntas/variações.")


if __name__ == "__main__":
    build_workbook()

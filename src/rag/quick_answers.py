import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
QUESTION_BANK_PATH = ROOT_DIR / "base_perguntas_respostas_ppgd.xlsx"
FAQ_PATH = ROOT_DIR / "static" / "FAQ.md"

DIRECT_THRESHOLD = 0.86
ASSIST_THRESHOLD = 0.70
SUGGEST_THRESHOLD = 0.58

STOPWORDS = {
    "a", "ao", "aos", "as", "ate", "com", "como", "da", "das", "de", "do",
    "dos", "e", "em", "eu", "me", "minha", "meu", "na", "nas", "no", "nos",
    "o", "os", "ou", "para", "por", "qual", "quais", "que", "sao", "se",
    "sobre", "um", "uma", "voce", "pode", "gostaria", "saber",
}

IMPORTANT_TERMS = {
    "credito", "creditos", "disciplina", "disciplinas", "prazo", "prazos",
    "defesa", "qualificacao", "suficiencia", "email", "telefone", "ramal",
    "instagram", "fomento", "dissertacao", "atividades", "complementares",
    "matricula", "aproveitamento", "lingua", "estrangeira",
}


@dataclass(frozen=True)
class QuickAnswer:
    question: str
    canonical_question: str
    answer: str
    source: str
    category: str
    intent: str
    response_type: str
    fallback_answer: str


@dataclass(frozen=True)
class QuickMatch:
    answer: QuickAnswer
    score: float
    mode: str


def _fold(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return without_accents.lower()


def _normalize(text: str) -> str:
    text = _fold(text)
    text = re.sub(r"[^a-z0-9@./:-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _stem(word: str) -> str:
    suffixes = ("ador", "adora", "acao", "ação", "ativo", "ativa", "mente", "s")
    for suffix in sorted(suffixes, key=len, reverse=True):
        if word.endswith(suffix) and len(word) - len(suffix) >= 4:
            return word[: -len(suffix)]
    return word


def _tokens(text: str) -> set[str]:
    raw_tokens = {word for word in _normalize(text).split() if len(word) >= 3 and word not in STOPWORDS}
    return {_stem(word) for word in raw_tokens}


def _clean_markdown(text: str) -> str:
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text or "")
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _score(query: str, candidate: str) -> float:
    normalized_query = _normalize(query)
    normalized_candidate = _normalize(candidate)
    if not normalized_query or not normalized_candidate:
        return 0.0
    if normalized_query == normalized_candidate:
        return 1.0

    seq_score = SequenceMatcher(None, normalized_query, normalized_candidate).ratio()
    query_tokens = _tokens(query)
    candidate_tokens = _tokens(candidate)
    if not query_tokens or not candidate_tokens:
        return seq_score

    query_important = query_tokens & IMPORTANT_TERMS
    candidate_important = candidate_tokens & IMPORTANT_TERMS
    if query_important and not (query_important & candidate_important):
        seq_score = min(seq_score, 0.45)

    overlap = len(query_tokens & candidate_tokens)
    containment = overlap / max(1, len(query_tokens))
    jaccard = overlap / max(1, len(query_tokens | candidate_tokens))
    return max(seq_score, containment * 0.9, jaccard)


def _shares_important_term(query: str, answer: QuickAnswer) -> bool:
    query_terms = _tokens(query) & IMPORTANT_TERMS
    answer_terms = (_tokens(answer.question) | _tokens(answer.canonical_question) | _tokens(answer.answer)) & IMPORTANT_TERMS
    return bool(query_terms and query_terms & answer_terms)


def _parse_faq_entries() -> list[QuickAnswer]:
    if not FAQ_PATH.exists():
        return []

    text = FAQ_PATH.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^###\s+(.+)$", text, flags=re.MULTILINE))
    entries = []

    for index, match in enumerate(matches):
        question = re.sub(r"^\d+\.\s*", "", match.group(1)).strip()
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        source_match = re.search(r"Fonte[s]?:\s*(.+)", block, flags=re.IGNORECASE)
        source = source_match.group(1).strip() if source_match else "static/FAQ.md"
        answer = re.split(r"Fonte[s]?:", block, maxsplit=1, flags=re.IGNORECASE)[0]
        answer = _clean_markdown(" ".join(line.strip() for line in answer.splitlines() if line.strip()))
        if question and answer:
            entries.append(
                QuickAnswer(
                    question=question,
                    canonical_question=question,
                    answer=answer,
                    source=source,
                    category="faq",
                    intent=_normalize(question).replace(" ", "_")[:48],
                    response_type="faq",
                    fallback_answer="Desculpe, não tenho informações suficientes para responder a essa pergunta.",
                )
            )
    return entries


def _load_from_workbook() -> list[QuickAnswer]:
    if not QUESTION_BANK_PATH.exists():
        return []

    workbook = load_workbook(QUESTION_BANK_PATH, read_only=True, data_only=True)
    sheet = workbook["Perguntas_Respostas"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: position for position, name in enumerate(headers)}
    entries = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[index["pergunta_variacao"]]:
            continue
        entries.append(
            QuickAnswer(
                question=str(row[index["pergunta_variacao"]]),
                canonical_question=str(row[index["pergunta_base"]]),
                answer=_clean_markdown(str(row[index["resposta_esperada"]])),
                source=str(row[index["fonte"]]),
                category=str(row[index["categoria"]]),
                intent=str(row[index["intencao"]]),
                response_type=str(row[index["tipo_resposta"]]),
                fallback_answer=_clean_markdown(str(row[index["resposta_quando_nao_entender"]])),
            )
        )
    return entries


@lru_cache(maxsize=1)
def load_quick_answers() -> tuple[QuickAnswer, ...]:
    entries = _load_from_workbook()
    if not entries:
        entries = _parse_faq_entries()
    return tuple(entries)


def _direct_url_answer(question: str, context: str) -> str | None:
    if not re.search(r"\b(link|site|página|pagina|instagram|rede social|redes sociais|acesso|acessar|url)\b", question, re.I):
        return None
    if not context:
        return None

    urls = re.findall(r"https?://[^\s)>\"]+", context)
    if not urls:
        return None

    lowered_question = question.lower()
    question_terms = [w for w in re.findall(r"[\wçãõáéíóúâêô]+", _fold(lowered_question)) if len(w) >= 4]

    for match in re.finditer(r"https?://[^\s)>\"]+", context):
        url = match.group(0)
        if "instagram" in lowered_question and "instagram" not in url.lower():
            continue
        window_start = max(0, match.start() - 120)
        window = _fold(context[window_start:match.start()])
        if any(term in window for term in question_terms if term not in ("site", "pode", "passar")):
            return url.rstrip(".,;")

    for url in urls:
        if "instagram" in lowered_question and "instagram" not in url.lower():
            continue
        return url.rstrip(".,;")
    return urls[0].rstrip(".,;")


def _direct_contact_answer(question: str, context: str) -> str | None:
    if re.search(r"\b(e-?mail|email|correio eletrônico)\b", question, re.I):
        emails = re.findall(r"[\w.+-]+@[\w.-]+\.\w+", context or "")
        if emails:
            return f"O e-mail do mestrado é {emails[0].rstrip('.,;')}"

    if re.search(r"\b(telefone|ramal|ligar|contato)\b", question, re.I):
        phones = re.findall(r"(?:\(?\d{2}\)?\s*)?\d{4,5}[- ]?\d{4}(?:\s*/\s*\d{4})?", context or "")
        if phones:
            return f"O ramal de contato do mestrado é {phones[0].strip().rstrip('.,;')}"

    return None


def _direct_academic_answer(question: str, context: str) -> str | None:
    folded_question = _fold(question)
    folded_context = _fold(context)

    asks_credits = "credito" in folded_question or "creditos" in folded_question
    mentions_whole_program = any(term in folded_question for term in ["curso", "programa", "mestrado", "integraliza", "grade curricular"])
    asks_total = mentions_whole_program and any(term in folded_question for term in ["tenho que ter", "preciso ter", "total", "quantos"])
    has_total = "totalizam-se 33" in folded_context or "33 (trinta e tres) creditos" in folded_context

    if asks_credits and asks_total and has_total:
        return (
            "A grade curricular totaliza 33 (trinta e três) créditos: 9 de formação geral, "
            "6 de aprofundamento específico na linha de pesquisa, 6 de aprofundamento específico "
            "de livre escolha, 4 de imersão prático-institucional, 4 de discussão e disseminação "
            "do conhecimento e 4 de pesquisa/escrita acadêmica."
        )

    return None


def find_quick_match(query: str, threshold: float = SUGGEST_THRESHOLD) -> Optional[QuickMatch]:
    # OBS: este arquivo já teve, em algum momento, um bloco grande de atalhos
    # "hardcoded" por pergunta especifica (checando frases literais como
    # "michigan ecce", "docencia", "qualificacao escandinava" etc, cada um
    # devolvendo uma resposta pronta com score=1.0). Isso foi removido de
    # proposito: aqueles atalhos so reconheciam a frase EXATA das perguntas
    # do dataset de teste, entao qualquer aluno real perguntando a mesma
    # coisa com outras palavras nunca cairia neles - o sistema parecia bom
    # no RAGAS mas nao generalizava. A partir daqui, so existe o fuzzy-match
    # generico contra a planilha/FAQ (por similaridade), que e' o mecanismo
    # legitimo de atalho. Tudo o mais deve passar pelo RAG de verdade.
    best_answer = None
    best_score = 0.0

    for answer in load_quick_answers():
        score = max(_score(query, answer.question), _score(query, answer.canonical_question) * 0.96)
        if score > best_score:
            best_score = score
            best_answer = answer

    if not best_answer or best_score < threshold:
        return None

    if best_score >= DIRECT_THRESHOLD:
        mode = "direct"
    elif best_score >= ASSIST_THRESHOLD or (best_score >= 0.58 and _shares_important_term(query, best_answer)):
        mode = "assist"
    else:
        mode = "suggest"

    return QuickMatch(answer=best_answer, score=round(best_score, 3), mode=mode)


def quick_context(match: QuickMatch) -> str:
    answer = match.answer
    return (
        "Resposta provável vinda da base rápida de FAQ/planilha:\n"
        f"Pergunta semelhante: {answer.canonical_question}\n"
        f"Resposta esperada: {answer.answer}\n"
        f"Fonte: {answer.source}\n"
        f"Similaridade: {match.score}"
    )


_REFUSAL_PATTERNS = (
    "nao tenho informacoes suficientes",
    "nao tenho informacoes o suficiente",
    "nao possuo informacoes suficientes",
    "nao encontrei essa informacao",
    "nao encontrei informacoes",
    "nao ha informacoes suficientes",
    "nao consta essa informacao",
    "nao consta informacao",
    "nao foi possivel encontrar",
    "nao tenho essa informacao",
    "nao tenho dados suficientes",
    "nao disponho de informacoes",
    "fora do escopo",
    "nao esta contemplado",
    "nao consigo responder",
    "nao sei informar",
    "sem informacoes suficientes",
)


def is_insufficient_answer(text: str) -> bool:
    folded = _fold(text or "")
    return any(pattern in folded for pattern in _REFUSAL_PATTERNS)


def sources_for_quick_match(match: QuickMatch) -> list[dict]:
    answer = match.answer
    return [
        {
            "source": f"Base rápida: {answer.source}",
            "excerpt": f"Pergunta semelhante: {answer.canonical_question} (similaridade {match.score})",
        }
    ]
